#! python3

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import shelve, os, time, sys
import openpyxl
import random
# housekeeping login and such
os.chdir('C:\\Programs')
browser = webdriver.Chrome('C:\\Program Files (x86)\\Python36-32\\chromedriver-Windows')
browser.get('http://apexvs.com')
elementz = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.NAME, 'ctl00$ContentPlaceHolder1$loginUsernameTextBox')))
login = browser.find_element_by_name('ctl00$ContentPlaceHolder1$loginUsernameTextBox')
login.send_keys('login')
passwd= browser.find_element_by_name('ctl00$ContentPlaceHolder1$passwordTextBox')
passwd.send_keys('pass')
passwd.submit()
time.sleep(3) #for loading time

#todays date in a format consistent with apex 

import datetime
dt=datetime.datetime.now()
if dt.month == 1:
    month = 'Jan'
elif dt.month == 2:
    month = 'Feb'
elif dt.month == 3:
    month = 'Mar'
elif dt.month == 4:
    month = 'Apr'
elif dt.month == 5:
    month = 'May'
elif dt.month == 6:
    month = 'Jun'
elif dt.month == 7:
    month = 'Jul'
elif dt.month == 8:
    month = 'Aug'
elif dt.month == 9:
    month = 'Sep'
elif dt.month == 10:
    month = 'Oct'
elif dt.month == 11:
    month = 'Nov'
elif dt.month == 12:
    month = 'Dec'
toDay=str(dt.day)
if len(str(dt.day))==1:
    toDay='0'+str(dt.day)
    

todayDate = toDay + " " + month + " " + str(dt.year)
print (todayDate)

wb = openpyxl.load_workbook('MasterActivity.xlsx')
sheet = wb.active
maxcol= sheet.max_column +1
sheet.cell(row=1, column=maxcol).value = todayDate
wb.save('MasterActivity.xlsx')

shelfFile = shelve.open('classes')
shelfFile2 = shelve.open('url')



studentlist = list(shelfFile2.keys())
random.shuffle(studentlist)

for student in studentlist:
    daycount = 0
    monthcount = 0
    yearcount = 0
    os.chdir('C:\\Programs')
    try:
        url = shelfFile2[student] #gets the main url for the student
        classurls = shelfFile[url] #takes the main url and gives a list of the class urls
    
  
        for y in classurls:  #loops through the class urls in the list
        
            print(y)
            browser.get(y)
            cell = []
            celltext = []
            
            for x in range(100):   #populates the cell list with all completed assignments
                element = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.ID, 'reportGrid_cell_0_4')))
                cell.append(browser.find_elements_by_id('reportGrid_cell_%s_4' % x))
            for unit in cell:  #converts webelement to date
                for point in unit:
                    celltext.append(point.text)
            for z in celltext:   #checks the dates and adds to the counters
               
                if (month + ' ' + str(dt.year)) in z:
                    monthcount += 1
                    
                if z == todayDate:
                    daycount += 1
                    
                if str(dt.year) in z:
                    yearcount +=1
    except Exception as err:
        print('there was an error: %s' % (err))    
        
    print(student)  
    print('Total assignments today: ' +str(daycount))

    
    #exports to a spreadsheet
    
    
        
##    shelfFile3['2017'] += 1 #shelfFile[2017] needs to be defined previously in idle
    os.chdir('C:\\Programs\\Activity')

    if student + '.xlsx' not in os.listdir('C:\\Programs\\Activity'): # this block makes each individual student excel
        wb = openpyxl.Workbook()
        wb.save(student + '.xlsx')
    wb = openpyxl.load_workbook(student + '.xlsx')
    sheet = wb.active
    maxcol= sheet.max_column +1
    sheet.cell(row=1, column=maxcol).value = todayDate #make '2017' into student if you want to switch back to individual counts
    sheet.cell(row=2, column=maxcol).value = daycount

    wb.save(student + '.xlsx')


    totalfile = open(todayDate+'.txt', 'a')  #saves a text file of the output
    totalfile.write(student+'\n')
    totalfile.write('Total assignments today: ' +str(daycount)+'\n')

    totalfile.close()
    os.chdir('C:\\Programs')

    wb = openpyxl.load_workbook('MasterActivity.xlsx')
    sheet = wb.active
    

    oldstudent=False
    maxcol= sheet.max_column
    maxrow= sheet.max_row+1
    for i in range(1,maxrow):
        if student == sheet.cell(row=i, column=1).value:
            sheet.cell(row=i, column=maxcol).value = daycount
            oldstudent=True
    if oldstudent==False:
        sheet.cell(row=maxrow, column=1).value = student
        sheet.cell(row=maxrow, column=maxcol).value = daycount
    print(oldstudent)

##    print(oldstudent)
    wb.save('MasterActivity.xlsx')


    
os.system("TASKKILL /F /IM chromedriver.exe")
os.system("taskkill /f /im  chrome.exe")
shelfFile.close()
shelfFile2.close()

